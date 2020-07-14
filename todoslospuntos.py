#Este paquete es para guardar un archivo csv
import pandas as pd

#Este paquete es el de para transformar coordenadas
import pyproj
from pyproj import CRS

#Este paquete es el que te encuentra si un punto cae dentro de un polígono o no
from shapely.geometry import Point
from shapely.geometry.polygon import Polygon

#Este pos es el de para leer el excel
from openpyxl import load_workbook

#Este pos es el de para leer el .shp
import shapefile

#Modulo: Leer las coordenadas de las tiendas

#Este es pa leer las coordenadas de las tiendas
wb = load_workbook('coordenadas.xlsx')
sheet = wb.worksheets[0]

#En el siguiente 'for' las metes en la lista 'coordenadas'
coordenadas = []

for row in sheet.iter_rows(min_row=2,values_only=True):
	xy = [ row[0] , row[1]]
	if xy not in coordenadas:
		coordenadas.append(xy)


#Modulo: Hacer el transformador de coordenadas

#Este pos ya viene explicado en la bitácora

crs = CRS.from_epsg(6372)

crs2 = CRS.from_epsg(4326)

proj = pyproj.transformer.Transformer.from_crs(crs, crs2)

#Modulo: Leer el shapefile para sacar los polígonos

shape = shapefile.Reader('09a.shp')

#Submodulo: En esta parte es donde ocurre la magia

#La lista puntos es la que va a tener a los puntos que sí estén dentro de determinado polígono
puntos = []
#La lista coort una vez más es auxiliar
coort = []
#La lista poli es donde estarán los puntos que sí estén dentro de determinado polígono y aparte
#tendrán al lado el código correspondiente al AGEB dentro del que cáen
poli = [[]]
#Las siguientes 2 variables namas son contadores, uno para la shape y otro para las coordenadas.
#El pedo es que el for se va de shape en shape, osea de polígono en polígono, pero la variable shape 
#como tal no es un número, entonces no se puede usar como índice para agregar algo a una lista, entonces
#literalmente pongo un número (en este caso j) que acompañe a cada shape y sea como su índice
#y lo mismo aplica para k porque cada par de coordenadas de una tienda que está dentro de la varible 
#'coordenadas' que se creó más arriba cuando se leyó el excel tampoco viene siendo un número según yo
#entonces k va a acompañar a cada punto de esa lista y va a ser su índice.
j = 0
k = 0
#shape es cada polígono
for shape in shape.shapeRecords():
    #En el siguiente for nada más se transforman y guardan las coordenadas de cada polígono en la variable
    #auxiliar 'coort'
	for i in shape.shape.points[:]:
		#Se transforman las coordenadas
		t = proj.transform(i[1],i[0])
		x = t[0]
		y = t[1]
		xy = [x,y]
        #coort trae los puntos de cada polígono y luego se reinicia
		coort.append(xy)
	#Se crea el polígono con las coordenadas que se acaban de obtener en el for anterior
    polygon = Polygon(coort)
    #En este for se agarran las coordenadas que salieron del excel de las tiendas y se busca si 
    #alguna de ellas cae dentro del polígono que se acaba de agarrar en el for pasado
	for item in coordenadas:
		point = Point(item[0], item[1])
        #Con el if checamos si si esta dentro
		if polygon.contains(point):
			x3 = item[0]
			y3 = item[1]
			mm = [x3,y3]
            #puntos trae los puntos que sí cayeron dentro del polígono
			puntos.append(mm)
            #Aparte, si la coordenada 'k' si cayó dentro, le digo que la quite de la lista para que 
            #la siguiente vez que se corra este for ya no esté ahí ocupando espacio porque ya sabemos que 
            #ya cayó dentro de un polígono
			coordenadas.pop(k)
        #Le incremento a la variable k porque sigue la siguiente coordenada
		k = k + 1
    #Ya que se acabó el for anterior, acomodo los datos en la lista ab de manera que quede en la izquierda
    #las coordenadas y en la derecha los datos del polígono
	ab = [puntos,shape.record]
    #Le inserto a otra lista que cree afuera de todo esto el ab, en el índice j, para eso fue 
    #que cree la variable j afuera de los fors, para que me ayude a meter los datos del polígono 'j' 
    #en la posición 'j' de la lista
	poli.insert(j,ab)
    #Reinicio k para la siguiente iteración
	k = 0
    #Reincio las listas para la siguiente iteración
	coort = []
	puntos = []
    #Le sumo 1 a j para indicar que sigue el polígono j+1
	j = j+1

#El print te dice cuantos pares de coordenadas quedaron en la variable coordenadas, osea cuantas coordenadas
#no cayeron dentro de ningun polígono
print(len(coordenadas))

#Modulo: hacer el csv

#Con eso literalmente nada más se hace el csv con toda la información que nos quedó en la lista 'poli'
#Lo tengo comentado porque cada que lo corras y no este comentado te va a crear un archivo nuevo y pues yo
#ya tengo ese archivo entonces ya no necesito correr esta parte.
#df = pd.DataFrame(poli)
#df.to_csv('sinindices.csv', index=False, header=False)

#Fin

