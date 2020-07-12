from openpyxl import load_workbook
import numpy as np
import pandas as pd

wb = load_workbook('sales_sample.xlsx')
sheet = wb.worksheets[0]

#
#Tamaño del excel:

#row_count = sheet.max_row
#column_count = sheet.max_column
#print(f'Número de filas: {row_count}')
#print(f'Número de columnas: {column_count}')

#
#Ver los títulos:

titulos = []

for row in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
	titulos += row

#print(titulos)

#
#Obtener las puras coordenadas de cada tienda:

#indicelat = titulos.index('Lat')
#indicelong = titulos.index('Long')
#print(f'Indice de Lat: {indicelat}')
#print(f'Indice de Lon: {indicelong}')

#coordenadas = []

#for row in sheet.iter_rows(min_row=2,values_only=True):
#	latitud = row[indicelat]
#	longitud = row[indicelong]
#	xy = [latitud,longitud]
#	if xy not in coordenadas:
#		coordenadas.append(xy)

#print(coordenadas)

#
#Separar por productos los datos de Kellogs

productos = []
indiceprod = titulos.index('Product_Code')
indicedes = titulos.index('Product_Description')

for row in sheet.iter_rows(values_only=True):
	codigo = row[indiceprod]
	des = row[indicedes]
	xy = [codigo, des]
	if xy not in productos:
		productos.append(xy)

tienditas = []

for row in productos:
	codigo = row[0]
	nombre = row[1]
	for row in sheet.iter_rows(values_only=True):
		c = row[indiceprod]
		if c == codigo or c == 'Product_Code':
			tienditas.append(row)
	m = np.matrix(tienditas)
	df = pd.DataFrame(data=m)
	df.to_csv(nombre + '.csv', sep=' ', header=False, index=False)
	tienditas.clear()
