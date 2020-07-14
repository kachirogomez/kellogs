import matplotlib.pyplot as plt

import folium
from folium.plugins import FastMarkerCluster

import pyproj
from pyproj import CRS

from shapely.geometry import Point
from shapely.geometry.polygon import Polygon

from openpyxl import load_workbook

#Transformador de coordenadas

crs = CRS.from_epsg(6372)

crs2 = CRS.from_epsg(4326)

proj = pyproj.transformer.Transformer.from_crs(crs, crs2)

#Leer el shapefile
import shapefile
shape = shapefile.Reader('09a.shp')
shape = shape.shapeRecords()
#Agarrar el primer polígono
shape0 = shape[0]
#Agarrar sus puntos
coord = shape0.shape.points[:]
#Inicarlizar listas vacías para usar en el for
coord2 = []
eq = []
ye = []

#Esta namas es para iniciar una figura en matplotlib
plt.figure()
#For para transformar las coordenadas y añadirlas a la lista coord2
for i in range(len(coord)):
	t = proj.transform(coord[i][1],coord[i][0])
	x1 = t[0]
	y1 = t[1]
	eq.append(y1)
	ye.append(x1)
	xy = [x1,y1]
	coord2.append(xy)

#Aqui namas le dices al matplot que te ponga el poligono en la lista
plt.plot(eq,ye)

print('Estos son los puntos del polígono')
print(coord2)

#Aqui le dices al shapely que haga el polígono con las coordenadas coord2
polygon = Polygon(coord2)

#Este es pa leer las coordenadas de las tiendas
wb = load_workbook('coordenadas.xlsx')
sheet = wb.worksheets[0]

#En el siguiente for las metes en la lista 'coordenadas'
coordenadas = []

for row in sheet.iter_rows(min_row=2,values_only=True):
	xy = [ row[0] , row[1]]
	if xy not in coordenadas:
		coordenadas.append(xy)

#En el siguiente for checas si los puntos están en el polígono
puntos = [[]]

for item in coordenadas:
	point = Point(item[0], item[1])
	if polygon.contains(point):
		x3 = item[0]
		y3 = item[1]
		plt.plot([y3],[x3],marker='o', markersize=3, color="red")
		mm = [x3,y3]
		puntos.append(mm)

print('Estos son los puntos adentro del polígono')
print(puntos)
#Y por último le dices a matplotlib que te enseñe todo
plt.show()